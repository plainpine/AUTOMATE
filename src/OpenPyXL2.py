import openpyxl
wb = openpyxl.load_workbook('example.xlsx')
wb.sheetnames # ワークブックに含まれるシート名一覧
print(wb.sheetnames)
sheet = wb['Sheet3'] # ワークブックからシートを取得
sheet
print(sheet)
type(sheet)
print(type(sheet))
sheet.title # シート名を文字列として取得
print(sheet.title)
anotherSheet = wb.active # アクティブなシートを取得
anotherSheet
print(anotherSheet)
