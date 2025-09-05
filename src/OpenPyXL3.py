import openpyxl
wb = openpyxl.load_workbook('example.xlsx')
sheet = wb['Sheet1'] # ワークブックからシートを取得
print(sheet)
sheet['A1'] # シートからセルを取得
print(sheet['A1'] )
sheet['A1'].value # セルの値を取得
print(sheet['A1'].value)
c = sheet['B1'] # 別のセルを取得
c.value
print(c.value)
# セルの列、行、値を取得
'Row %s, Column %s is %s' % (c.row, c.column, c.value)
print('Row %s, Column %s is %s' % (c.row, c.column, c.value))
'Cell %s is %s' % (c.coordinate, c.value)
print('Cell %s is %s' % (c.coordinate, c.value))
sheet['C1'].value
print(sheet['C1'].value)

sheet.cell(row=1, column=2)
print(sheet.cell(row=1, column=2))

sheet.cell(row=1, column=2).value
print(sheet.cell(row=1, column=2).value)
for i in range(1, 8, 2): # 行を１つおｌきに調べる
    print(i, sheet.cell(row=i, column=2).value)
