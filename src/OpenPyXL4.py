import openpyxl
wb = openpyxl.load_workbook('example.xlsx')
sheet = wb['Sheet1']
sheet.max_row # 最大行数を取得
print(sheet.max_row)

sheet.max_column # 最大列数を取得
print(sheet.max_column)
