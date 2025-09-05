import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
get_column_letter(1) # 列番号を文字に変換
print(get_column_letter(1))

get_column_letter(2)
print(get_column_letter(2))

get_column_letter(27)
print(get_column_letter(27))

get_column_letter(900)
print(get_column_letter(900))

wb = openpyxl.load_workbook('example.xlsx')
sheet = wb['Sheet1']
get_column_letter(sheet.max_column)
print(get_column_letter(sheet.max_column))

column_index_from_string('A') # 列名を列番号に変換
print(column_index_from_string('A'))

column_index_from_string('AA')
print(column_index_from_string('AA'))

