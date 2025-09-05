import openpyxl
wb = openpyxl.Workbook()
wb.sheetnames
print(wb.sheetnames)

wb.create_sheet() # 新しいシート作成
wb.sheetnames
print(wb.sheetnames)

# 新しいシートをインデックス0に作成
wb.create_sheet(index=0, title='First Sheet')
wb.sheetnames
print(wb.sheetnames)

wb.create_sheet(index=2, title='Middle Sheet')
wb.sheetnames
print(wb.sheetnames)

del wb['Middle Sheet']
del wb['Sheet1']
wb.sheetnames
print(wb.sheetnames)
