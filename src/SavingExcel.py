import openpyxl
wb = openpyxl.Workbook() # 空白のワークブックを作成
wb.sheetnames # 最初から１つのシートがある
print(wb.sheetnames)

sheet = wb.active
sheet.title
print(sheet.title)

sheet.title = 'Spam Bacon Eggs Sheet' # シート名を変更
wb.sheetnames
print(wb.sheetnames)
print(sheet.title)
