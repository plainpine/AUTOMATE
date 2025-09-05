"""
行と列の入れ替え
"""

def main():
    import sys
    import openpyxl

    # Get argument from commandline
    file = ''.join(sys.argv[1])

    # Open workbooks
    readwb = openpyxl.load_workbook(file)
    readsheet = readwb.active

    writewb = openpyxl.Workbook()
    writesheet = writewb.active

    # Read readwb and transpose into writewb
    for rowNum in range(1, readsheet.max_row + 1):
        for colNum in range(1, readsheet.max_column + 1):
                # Invert columns and rows
                writesheet.cell(row=colNum, column=rowNum).value = readsheet.cell(row=rowNum, column=colNum).value

    # Save workbook
    writewb.save("updated" + file)

if __name__ == '__main__':
    main()
