"""
Multiplication table
コマンドラインから数Nを受け取り、
ExcelスプレッドシートにN×Nの乗算表を作成する
プログラムmultiplicationTable.pyを作成しなさい。
1行目とA列はラベルに使用し、太字にします。
"""

def main():
    import sys
    import openpyxl
    from openpyxl.styles import Font

    # Get argument from commandline
    size = int(''.join(sys.argv[1]))

    # Build table
    wb = openpyxl.Workbook()
    sheet = wb.active
    row = []
    for i in range(1, size + 1):
        row.append(i)
    column = row

    # Make labels
    for element in row:
        sheet.cell(row=1, column=element + 1).value = element
        sheet.cell(row=element + 1, column=1).value = element
    col = sheet.column_dimensions['A']
    col.font = Font(bold=True)
    ro = sheet.row_dimensions[1]
    ro.font = Font(bold=True)

    # Calculate table values
    for element in row:
        for element2 in column:
            sheet.cell(row=element + 1, column=element2 + 1).value = element * element2

    # Save table
    wb.save("multTable.xlsx")

if __name__ == '__main__':
    main()
