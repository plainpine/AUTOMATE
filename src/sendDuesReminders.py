#! python3
# sendDuesReminders.py - スプレッドシートの支払い状況に基づきメールを送信

import openpyxl, smtplib, sys
from mails import send_email

# スプレッドシートを開き最近の支払い状況を取得
wb = openpyxl.load_workbook('duesRecords.xlsx')  # ❶
sheet = wb['Sheet1'] # ❷

last_col = sheet.max_column  # ❸
latest_month = sheet.cell(row=1, column=last_col).value  # ❹

# 会員の支払い状況を調べる
unpaid_members = {}
for r in range(2, sheet.max_row + 1):        # ❶
    payment = sheet.cell(row=r, column=last_col).value # ❷
    if payment != 'paid':
        name = sheet.cell(row=r, column=1).value       # ❸
        email = sheet.cell(row=r, column=2).value      # ❹
        unpaid_members[name] = email  # ❺

# リマインダーメールを送信する
for name, email in unpaid_members.items():
    mail_title = f"{latest_month} dues unpaid."
    mail_text = f"Dear {name}, Records show that you have not paid dues for {latest_month}. \n Please make this payment as soon as possible. Thank you!"
    send_email(mail_title, mail_text, email)



