import openpyxl
wb = openpyxl.load_workbook('merged.xlsx')
sheet = wb.active
sheet.unmerge_cells('A1:D3') # セルのマージを解除
sheet.unmerge_cells('C5:D5')
wb.save('merged.xlsx')
